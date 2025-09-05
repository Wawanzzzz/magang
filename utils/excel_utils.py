# utils/excel_utils.py
import re
import shutil
import tempfile
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook

OUTPUT_EXCEL = Path("assets/output/ocr_results.xlsx")

# Struktur kolom final
HEADERS = [
    "Timestamp", "Source File",
    "nik", "nama", "tempat_lahir", "tanggal_lahir",
    "jenis_kelamin", "golongan_darah", "alamat", "rt", "rw",
    "kelurahan_atau_desa", "kecamatan", "agama",
    "status_perkawinan", "pekerjaan", "kewarganegaraan", "berlaku_hingga"
]

# -------------------------
# Excel utilities
# -------------------------
def _ensure_workbook():
    """Buat workbook baru jika belum ada, dan pastikan header sesuai."""
    if OUTPUT_EXCEL.exists():
        wb = load_workbook(OUTPUT_EXCEL)
        ws = wb.active
        return wb, ws
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)
        return wb, ws

def _save_wb_safely(wb, target: Path):
    """Simpan workbook ke tmp file lalu move untuk mengurangi permission error."""
    tmp = Path(tempfile.gettempdir()) / f"ocr_results_tmp_{int(datetime.now().timestamp())}.xlsx"
    wb.save(tmp)
    try:
        shutil.move(str(tmp), str(target))
    except PermissionError:
        i = 1
        while True:
            alt = target.with_name(f"{target.stem}_{i}{target.suffix}")
            if not alt.exists():
                shutil.move(str(tmp), str(alt))
                return alt
            i += 1
    return target

# -------------------------
# Parsers
# -------------------------
def _extract_tokens_from_detail(detail_path: Path):
    """
    Ambil semua token dari file detail OCR EasyOCR.
    Return list token urut: ['NIK', '6472...', 'Nama', 'JOHN DOE', ...]
    """
    txt = detail_path.read_text(encoding="utf-8", errors="ignore")
    tokens = []
    for m in re.finditer(r"Text:\s*'([^']+)'", txt):
        tokens.append(m.group(1).strip())
    return tokens

# --- helper normalisasi & koreksi typo ---
_TYPOS = {
    # common OCR mistakes -> corrected
    "jonis": "jenis",
    "tompautgl": "tempat tanggal",
    "tompautglahir": "tempat tanggal",
    "tompat": "tempat",
    "tmpt": "tempat",
    "namat": "nama",
    "stalus": "status",
    "porkawnan": "perkawinan",
    "pokorjaan": "pekerjaan",
    "kowarganegaraan": "kewarganegaraan",
    "kevdesa": "kelurahan desa",
    "kocamatan": "kecamatan",
    "aama": "agama",
    "borlaku": "berlaku",
    "borlaku hingoa": "berlaku hingga",
    "rtiaw": "rt rw",
}

FIELD_ALIASES = {
    "nik": ["nik"],
    "nama": ["nama", "namat"],
    "tempat_lahir": ["tempat", "lahir", "tmplahir", "tmpt", "tempat tanggal"],
    "tanggal_lahir": ["tanggal", "tgl"],
    "jenis_kelamin": ["jenis kelamin", "jenis", "jns kelamin"],
    "golongan_darah": ["gol darah", "goldarah", "gol"],
    "alamat": ["alamat"],
    "rt_rw": ["rt", "rw"],
    "kelurahan_atau_desa": ["kelurahan", "desa", "kel"],
    "kecamatan": ["kecamatan"],
    "agama": ["agama"],
    "status_perkawinan": ["status", "perkawinan"],
    "pekerjaan": ["pekerjaan"],
    "kewarganegaraan": ["kewarganegaraan", "wni", "wna"],
    "berlaku_hingga": ["berlaku", "berlaku hingga", "hingga"],
}

_JOB_KEYWORDS = [
    "PELAJAR", "MAHASISWA", "PNS", "PNS/TNI", "KARYAWAN", "WIRASWASTA",
    "BURUH", "PETANI", "PENSIUNAN", "SWASTA", "TNI", "POLRI", "PELAJAR/MAHASISWA"
]

def _apply_typo_corrections(s: str) -> str:
    s2 = s
    for bad, good in _TYPOS.items():
        # replace ignoring case
        s2 = re.sub(re.escape(bad), good, s2, flags=re.IGNORECASE)
    return s2

def _normalize_token(t: str) -> str:
    """Bersihkan token: ubah -/ jadi spasi, koreksi typo, hilangkan tanda baca kecuali alfanumerik/spasi."""
    if t is None:
        return ""
    # ubah beberapa separator jadi spasi supaya 'LAKI-LAKI' jadi 'LAKI LAKI'
    s = t.replace("-", " ").replace("/", " ").replace("\\", " ")
    s = _apply_typo_corrections(s)
    # remove punctuation except spaces and alnum
    s = re.sub(r"[^\w\s]", " ", s, flags=re.UNICODE)
    s = re.sub(r"\s+", " ", s).strip()
    return s.lower()

def _contains_alias(norm: str, field: str) -> bool:
    """Cek apakah token normalisasi mengandung salah satu alias field (kata lengkap)."""
    for a in FIELD_ALIASES.get(field, []):
        pattern = r"\b" + re.escape(a.lower()) + r"\b"
        if re.search(pattern, norm):
            return True
    return False

def _extract_value_after_last_alias(norm: str, field: str) -> str:
    """
    Jika token mengandung label dan nilai dalam satu string,
    cari posisi alias terakhir dan kembalikan sisa setelahnya.
    """
    last_pos = -1
    last_alias = None
    for a in FIELD_ALIASES.get(field, []):
        for m in re.finditer(r"\b" + re.escape(a.lower()) + r"\b", norm):
            if m.end() > last_pos:
                last_pos = m.end()
                last_alias = a
    if last_pos >= 0:
        value = norm[last_pos:].strip()
        # kalau sisanya masih kosong, bisa berarti nilai berada sebelumnya (rare)
        return value if value else None
    return None

def _standardize_gender(raw: str) -> str:
    if not raw:
        return ""
    r = raw.upper()
    if "LAKI" in r:
        return "Laki-laki"
    if "PEREMPUAN" in r or "WANITA" in r:
        return "Perempuan"
    return raw.title()

def _standardize_status(raw: str) -> str:
    if not raw:
        return ""
    r = raw.upper()
    if "BELUM" in r or "BUKAN" in r or "TIDAK" in r:
        if "KAWIN" in r:
            return "Belum Kawin"
    if "KAWIN" in r and "BELUM" not in r:
        return "Kawin"
    if "CERAI" in r:
        if "HIDUP" in r:
            return "Cerai Hidup"
        if "MATI" in r:
            return "Cerai Mati"
        return "Cerai"
    return raw.title()

def _clean_job(raw: str) -> str:
    if not raw:
        return ""
    ru = raw.upper()
    found = []
    for kw in _JOB_KEYWORDS:
        if kw in ru:
            # normalize form: PELAJAR + MAHASISWA -> Pelajar/Mahasiswa
            found.append(kw.replace("/", " / "))
    if found:
        # join unique and proper-cased (split slashes)
        parts = []
        for f in found:
            for p in f.split("/"):
                p = p.strip()
                if p and p not in parts:
                    parts.append(p)
        # Title-case each part and join with '/'
        return "/".join(p.title() for p in parts)
    # fallback: beri spasi sesuai input dan title-case
    cleaned = re.sub(r"\s+", " ", raw).strip()
    return cleaned.title()

def _standardize_kewarganegaraan(raw: str) -> str:
    if not raw:
        return ""
    r = raw.upper()
    if "WNI" in r:
        return "WNI"
    if "WNA" in r:
        return "WNA"
    # kalau ada kata Indonesia
    if "INDONES" in r:
        return "WNI"
    return raw.upper()

def _parse_tokens_to_fields(tokens):
    """Heuristik mapping token ke kolom KTP dengan toleransi typo dan label+value dalam satu token."""
    data = {h: "" for h in HEADERS[2:]}  # skip timestamp + source

    nik_re = re.compile(r"\b\d{15,18}\b")
    date_re = re.compile(r"\d{2}[-/]\d{2}[-/]\d{4}")
    rt_rw_re = re.compile(r"(\d{1,3})\s*(?:/|\s)\s*(\d{1,3})")

    i = 0
    while i < len(tokens):
        t = tokens[i].strip()
        t_norm = _normalize_token(t)
        nxt = tokens[i+1].strip() if i+1 < len(tokens) else ""
        nxt_norm = _normalize_token(nxt)

        # --- CASE: label and value in same token (try extract for several fields) ---
        for field in ("status_perkawinan", "kewarganegaraan", "jenis_kelamin", "pekerjaan", "berlaku_hingga", "agama"):
            if _contains_alias(t_norm, field):
                val = _extract_value_after_last_alias(t_norm, field)
                if val:
                    # assign based on field-specific standardization
                    if field == "jenis_kelamin":
                        data["jenis_kelamin"] = _standardize_gender(val)
                    elif field == "status_perkawinan":
                        data["status_perkawinan"] = _standardize_status(val)
                    elif field == "pekerjaan":
                        data["pekerjaan"] = _clean_job(val)
                    elif field == "kewarganegaraan":
                        data["kewarganegaraan"] = _standardize_kewarganegaraan(val)
                    elif field == "berlaku_hingga":
                        if "seumur" in val.lower():
                            data["berlaku_hingga"] = "Seumur Hidup"
                        else:
                            data["berlaku_hingga"] = val.title()
                    elif field == "agama":
                        data["agama"] = val.title()
                # even if no value in same token, we still continue to general parsing below
        # --- NIK ---
        if _contains_alias(t_norm, "nik") and nik_re.match(nxt):
            data["nik"] = nxt
            i += 1
            i += 1
            continue
        elif nik_re.match(t) and not data["nik"]:
            data["nik"] = t
            i += 1
            continue

        # --- Nama ---
        if _contains_alias(t_norm, "nama"):
            # if token contains both label+value (e.g. "Nama John Doe")
            val = _extract_value_after_last_alias(t_norm, "nama")
            if val:
                data["nama"] = val.title()
            else:
                data["nama"] = nxt.title()
                i += 1
            i += 1
            continue

        # --- Tempat & Tanggal Lahir ---
        if _contains_alias(t_norm, "tempat_lahir") or _contains_alias(t_norm, "tanggal_lahir"):
            # prefer pattern where next-next token adalah tanggal
            if i+2 < len(tokens) and date_re.match(tokens[i+2]):
                data["tempat_lahir"] = tokens[i+1].title()
                data["tanggal_lahir"] = tokens[i+2]
                i += 3
                continue
            # if next token contains date
            if date_re.search(nxt):
                # split mungkin "BATULICIN,KAB ... 26-04-2004" but usually date alone
                if " " in nxt:
                    parts = nxt.split()
                    if date_re.search(parts[-1]):
                        data["tempat_lahir"] = " ".join(parts[:-1]).title()
                        data["tanggal_lahir"] = parts[-1]
                    else:
                        data["tanggal_lahir"] = nxt
                else:
                    data["tanggal_lahir"] = nxt
                i += 1
                continue
            # else take next as tempat lahir
            if nxt:
                data["tempat_lahir"] = nxt.title()
                i += 1
                continue

        # --- Jenis Kelamin ---
        # if token contains label, get next or earlier-extracted same-token value
        if _contains_alias(t_norm, "jenis_kelamin"):
            # if current token contained value extracted earlier, it is already set via same-token logic
            if not data["jenis_kelamin"]:
                # check next token
                if nxt:
                    data["jenis_kelamin"] = _standardize_gender(nxt)
                    i += 1
            i += 1
            continue

        # --- Golongan Darah ---
        if _contains_alias(t_norm, "golongan_darah"):
            val = _extract_value_after_last_alias(t_norm, "golongan_darah")
            if val:
                data["golongan_darah"] = val.upper()
            else:
                data["golongan_darah"] = nxt.upper()
                i += 1
            i += 1
            continue

        # --- Alamat ---
        if _contains_alias(t_norm, "alamat"):
            data["alamat"] = nxt.title() if nxt else ""
            i += 1
            continue

        # --- RT/RW ---
        if _contains_alias(t_norm, "rt_rw") or rt_rw_re.search(nxt):
            # prefer to parse next token "005/003"
            m = rt_rw_re.search(nxt)
            if m:
                data["rt"], data["rw"] = m.group(1), m.group(2)
                i += 1
                continue
            # maybe current token contains 'rt 005 rw 003' compressed
            m2 = rt_rw_re.search(t_norm)
            if m2:
                data["rt"], data["rw"] = m2.group(1), m2.group(2)
            i += 1
            continue

        # --- Kelurahan / Desa ---
        if _contains_alias(t_norm, "kelurahan_atau_desa"):
            val = _extract_value_after_last_alias(t_norm, "kelurahan_atau_desa")
            if val:
                data["kelurahan_atau_desa"] = val.title()
            else:
                data["kelurahan_atau_desa"] = nxt.title()
                i += 1
            i += 1
            continue

        # --- Kecamatan ---
        if _contains_alias(t_norm, "kecamatan"):
            val = _extract_value_after_last_alias(t_norm, "kecamatan")
            if val:
                data["kecamatan"] = val.title()
            else:
                data["kecamatan"] = nxt.title()
                i += 1
            i += 1
            continue

        # --- Agama ---
        if _contains_alias(t_norm, "agama"):
            val = _extract_value_after_last_alias(t_norm, "agama")
            if val:
                data["agama"] = val.title()
            else:
                data["agama"] = nxt.title()
                i += 1
            i += 1
            continue

        # --- Status Perkawinan ---
        if _contains_alias(t_norm, "status_perkawinan"):
            # check same-token extraction already done; if not, get next
            if not data["status_perkawinan"]:
                val = _extract_value_after_last_alias(t_norm, "status_perkawinan")
                if val:
                    data["status_perkawinan"] = _standardize_status(val)
                elif nxt:
                    data["status_perkawinan"] = _standardize_status(nxt)
                    i += 1
            i += 1
            continue

        # --- Pekerjaan ---
        if _contains_alias(t_norm, "pekerjaan"):
            # prefer same-token extraction
            val = _extract_value_after_last_alias(t_norm, "pekerjaan")
            if val:
                data["pekerjaan"] = _clean_job(val)
            elif nxt:
                data["pekerjaan"] = _clean_job(nxt)
                i += 1
            i += 1
            continue

               # --- Kewarganegaraan ---
        if _contains_alias(t_norm, "kewarganegaraan"):
            val = _extract_value_after_last_alias(t_norm, "kewarganegaraan") or nxt
            standardized = _standardize_kewarganegaraan(val)
            if standardized in ("WNI", "WNA"):
                data["kewarganegaraan"] = standardized
            else:
                data["kewarganegaraan"] = ""
            i += 1
            continue

        # --- Berlaku Hingga ---
        if _contains_alias(t_norm, "berlaku_hingga"):
            data["berlaku_hingga"] = "Seumur Hidup"
            i += 1
            continue


        # fallback increment
        i += 1

    return data

# -------------------------
# Public function
# -------------------------
def save_to_excel_structured(detail_path: Path, source_file: str):
    """
    Simpan hasil OCR detail ke Excel.
    detail_path: path ke file *_detail.txt
    source_file: nama file gambar asli
    """
    wb, ws = _ensure_workbook()
    tokens = _extract_tokens_from_detail(detail_path)
    data = _parse_tokens_to_fields(tokens)

    row = [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), source_file] \
        + [data.get(h, "") for h in HEADERS[2:]]

    ws.append(row)
    _save_wb_safely(wb, OUTPUT_EXCEL)
    return OUTPUT_EXCEL